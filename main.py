import openpyxl

path = 'data/data_set.xlsx'
workbook = openpyxl.load_workbook(path)
sheet = workbook.active

#calculating avg of 21 days
summ = 0
ctr = 0
for value in sheet['B']:
    if ctr == 0:
        ctr = 1
        continue
    summ += int(value.value)
    ctr += 1
    if ctr > 21:
        break
print ("The average of 21 days is:\t",(summ/21))

#calculating avg of 50 days close price
ctr = 0
for value in sheet['B']:
    if ctr <= 21:
        ctr += 1
        continue
    summ += int(value.value)
    ctr += 1
    if ctr > 50:
        break
print ("The average of 50 days is:\t",(summ/50))


print ("-"*70)

#Give a data set which has all the open=high values.
allObj = sheet['A1':('E' + str(sheet.max_row))]
openHighLowObj = sheet['C1':('E' + str(sheet.max_row))]
openHighList = []
openLowList = []
for open, close, low in openHighLowObj:
    if (open.value == close.value):
        openHighList.append(open.value)
    if (open.value == low.value):
        openLowList.append(open.value)


for a,b,c,d,e in allObj:
    if (c.value in openHighList):
        print (a.value, b. value, c.value, d.value, e.value)

print ("-"*70)

for a,b,c,d,e in allObj:
    if (c.value in openLowList):
        print (a.value, b. value, c.value, d.value, e.value)




